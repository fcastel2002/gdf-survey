"""Excel workbook generator for GDF pump survey using openpyxl."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Sequence

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from gdf_survey.models import GdfSurveyResult

COLOR_HEADER_BG = "0F172A"       # Slate 900
COLOR_HEADER_FG = "FFFFFF"
COLOR_TITLE_BG = "1E293B"        # Slate 800
COLOR_SUBTITLE_BG = "334155"     # Slate 700
COLOR_ZEBRA_ODD = "F8FAFC"       # Slate 50
COLOR_ZEBRA_EVEN = "FFFFFF"
COLOR_BORDER = "CBD5E1"          # Slate 300
COLOR_YES_BG = "DCFCE7"          # Green 100
COLOR_YES_FG = "166534"          # Green 800
COLOR_NO_BG = "F1F5F9"           # Slate 100
COLOR_NO_FG = "94A3B8"           # Slate 400

CONTROLLER_COLORS = {
    "Tipo A": ("EFF6FF", "1E40AF"),      # Blue
    "Tipo B": ("FAF5FF", "6B21A8"),      # Purple
    "PLC": ("F0FDF4", "15803D"),         # Emerald
    "RTU": ("FFFBEB", "B45309"),         # Amber
    "Modulo IO": ("FEF2F2", "991B1B"),    # Red
}
BRAND_COLORS = CONTROLLER_COLORS


def _thin_border() -> Border:
    thin = Side(border_style="thin", color=COLOR_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_excel_survey(
    results: Sequence[GdfSurveyResult],
    output_path: str | Path,
) -> Path:
    """Generate a formatted multi-sheet Excel workbook with equipment survey results."""
    target_path = Path(output_path).resolve()
    target_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    # Summary sheet if multiple screens are surveyed
    if len(results) > 1:
        _build_summary_sheet(wb, results)

    # Consolidated pump sheet for each GDF result
    for res in results:
        _build_pumps_sheet(wb, res)

    wb.save(str(target_path))
    return target_path


def _build_summary_sheet(wb: openpyxl.Workbook, results: Sequence[GdfSurveyResult]) -> None:
    ws = wb.create_sheet(title="General Summary")
    ws.views.sheetView[0].showGridLines = True

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "SCADA EQUIPMENT GENERAL SURVEY - CONSOLIDATED SUMMARY"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    headers = [
        "Sheet / Display",
        "GDF File",
        "Surveyed Layer",
        "Surveyed Equipment",
        "Equipment with PT",
        "Controller Breakdown",
    ]
    ws.append([])
    ws.append(headers)
    ws.row_dimensions[3].height = 24

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    for r_idx, res in enumerate(results, start=4):
        brand_summary = ", ".join(f"{b}: {c}" for b, c in res.brand_counts.items())
        pt_count = sum(1 for p in res.pumps if p.has_pt == "1")
        row_vals = [
            res.sheet_name,
            res.gdf_path.name,
            res.layer_name,
            res.total_pumps,
            pt_count,
            brand_summary or "(No equipment)",
        ]
        ws.append(row_vals)
        ws.row_dimensions[r_idx].height = 20

        zebra = COLOR_ZEBRA_ODD if r_idx % 2 == 1 else COLOR_ZEBRA_EVEN
        for col_num, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=col_num)
            if isinstance(val, str):
                cell.data_type = "s"
            cell.border = _thin_border()
            cell.fill = PatternFill(start_color=zebra, end_color=zebra, fill_type="solid")
            cell.font = Font(name="Calibri", size=10)
            if col_num in (4, 5):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)


def _build_pumps_sheet(wb: openpyxl.Workbook, res: GdfSurveyResult) -> None:
    """Build the consolidated sheet with 1 row per pump showing all its configuration properties."""
    clean_title = re.sub(r'[\\/*?:\[\]]', "_", res.sheet_name)[:31]
    ws = wb.create_sheet(title=clean_title)
    ws.views.sheetView[0].showGridLines = True

    # 1. Main Banner
    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value = f"SCADA EQUIPMENT SURVEY - DISPLAY: {res.display_name}.gdf"
    title.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 2. Sub-banner
    ws.merge_cells("A2:L2")
    info = ws["A2"]
    brand_text = " | ".join(f"{b}: {c}" for b, c in res.brand_counts.items())
    pt_count = sum(1 for p in res.pumps if p.has_pt == "1")
    info.value = (
        f"Layer: {res.layer_name}   |   "
        f"Total Surveyed Items: {res.total_pumps}   |   "
        f"With PT Pressure: {pt_count}   |   "
        f"Controllers: {brand_text}"
    )
    info.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    info.fill = PatternFill(start_color=COLOR_SUBTITLE_BG, end_color=COLOR_SUBTITLE_BG, fill_type="solid")
    info.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # 3. Headers
    headers = [
        "No.",
        "Equipment",
        "Tag (<<pozo>>)",
        "Group (<<bat>>)",
        "Device (<<dispositivo>>)",
        "Controller",
        "Controller Type",
        "Has PT (<<tienept>>)",
        "Has TKE (<<tienetke>>)",
        "Has TKQ (<<tienetkq>>)",
        "Has SAM (<<tienesam>>)",
        "Is EXP (<<esexp>>)",
    ]

    ws.append([])
    ws.append(headers)
    ws.row_dimensions[4].height = 26

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_FG)
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    # 4. Data rows
    for r_idx, p in enumerate(res.pumps, start=5):
        row_vals = [
            p.pozo_index,
            p.well_id,
            p.pump_code,
            p.battery,
            p.device_name,
            p.controller_brand,
            p.controller_type,
            "YES" if p.has_pt == "1" else "NO",
            "YES" if p.has_tke == "1" else "NO",
            "YES" if p.has_tkq == "1" else "NO",
            "YES" if p.has_sam == "1" else "NO",
            "YES" if p.is_exp == "1" else "NO",
        ]
        ws.append(row_vals)
        ws.row_dimensions[r_idx].height = 20

        zebra = COLOR_ZEBRA_ODD if r_idx % 2 == 1 else COLOR_ZEBRA_EVEN
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=col_idx)
            if isinstance(val, str):
                cell.data_type = "s"
            cell.border = _thin_border()
            cell.font = Font(name="Calibri", size=9)
            cell.fill = PatternFill(start_color=zebra, end_color=zebra, fill_type="solid")

            # Column alignments
            if col_idx in (1, 2, 6, 7, 8, 9, 10, 11, 12):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Controller badge
            if col_idx == 6 and p.controller_brand in BRAND_COLORS:
                bg, fg = BRAND_COLORS[p.controller_brand]
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.font = Font(name="Calibri", size=9, bold=True, color=fg)

            # Yes / No flags styling
            if col_idx in (8, 9, 10, 11, 12):
                val = row_vals[col_idx - 1]
                if val == "YES":
                    cell.fill = PatternFill(start_color=COLOR_YES_BG, end_color=COLOR_YES_BG, fill_type="solid")
                    cell.font = Font(name="Calibri", size=9, bold=True, color=COLOR_YES_FG)
                else:
                    cell.fill = PatternFill(start_color=COLOR_NO_BG, end_color=COLOR_NO_BG, fill_type="solid")
                    cell.font = Font(name="Calibri", size=9, color=COLOR_NO_FG)

    end_col = get_column_letter(len(headers))
    last_row = 4 + len(res.pumps)
    ws.auto_filter.ref = f"A4:{end_col}{last_row}"
    ws.freeze_panes = "A5"

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        val_lens = [len(str(cell.value or "")) for cell in col[3:]]
        max_len = max(val_lens) if val_lens else 10
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 45), 12)
