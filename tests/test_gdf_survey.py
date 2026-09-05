"""Unit and integration tests for gdf_survey tool using synthetic fixtures."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from gdf_survey.cli import build_parser, main
from gdf_survey.excel_report import generate_excel_survey
from gdf_survey.extractor import extract_gdf_survey
from gdf_survey.html_report import generate_html_survey
from gdf_survey.models import classify_device_type


def test_classify_device_type() -> None:
    assert classify_device_type("DEV_01_CTRL_A")[0] == "Tipo A"
    assert classify_device_type("DEV_02_CTRL_B")[0] == "Tipo B"
    assert classify_device_type("DEV_03_IO_MOD")[0] == "Modulo IO"
    assert classify_device_type("PLC_LINE1")[0] == "PLC"
    assert classify_device_type("RTU_NORTH")[0] == "RTU"
    assert classify_device_type("UNKNOWN_DEVICE")[0] == "Desconocido"


def test_extract_survey_synthetic(synthetic_gdf: Path) -> None:
    res = extract_gdf_survey(synthetic_gdf, layer_target="1", sheet_name="Planta1")
    assert res.sheet_name == "Planta1"
    assert res.layer_name == "1-PROCESS"
    assert res.total_pumps == 1
    assert res.active_pumps == 1
    assert res.total_objects == 8
    assert res.brand_counts["Tipo A"] == 1

    p1 = res.pumps[0]
    assert p1.pozo_index == 1
    assert p1.well_id == "EQ_01"
    assert p1.pump_code == "TAG_101"
    assert p1.battery == "B-101"
    assert p1.controller_brand == "Tipo A"
    assert p1.has_pt == "1"
    assert p1.has_tke == "0"
    assert p1.has_tkq == "0"
    assert p1.has_sam == "0"
    assert p1.is_active is True

    assert "<<dispositivo>>" in res.custom_data_types
    assert "<<tienetke>>" in res.custom_data_types
    assert "<<tienept>>" in res.custom_data_types
    assert "<<pozo>>" in res.custom_data_types
    assert "<<bat>>" in res.custom_data_types


def test_excel_and_html_generation(synthetic_gdf: Path, tmp_path: Path) -> None:
    res = extract_gdf_survey(synthetic_gdf, layer_target="1", sheet_name="TestSheet")
    excel_out = tmp_path / "test_survey.xlsx"
    html_out = tmp_path / "test_survey.html"

    generate_excel_survey([res], excel_out)
    assert excel_out.exists()
    wb = openpyxl.load_workbook(excel_out)
    assert "TestSheet" in wb.sheetnames

    generate_html_survey([res], html_out)
    assert html_out.exists()
    content = html_out.read_text(encoding="utf-8")
    assert "TestSheet" in content
    assert "EQ_01" in content


def test_cli_execution(synthetic_gdf: Path, tmp_path: Path) -> None:
    excel_out = tmp_path / "cli_survey.xlsx"
    html_out = tmp_path / "cli_survey.html"

    exit_code = main([
        str(synthetic_gdf),
        "--name=Planta1",
        "--layer=1",
        f"--out-excel={excel_out}",
        f"--out-html={html_out}",
        "--quiet",
    ])
    assert exit_code == 0
    assert excel_out.exists()
    assert html_out.exists()


def test_cli_output_dir_options(synthetic_gdf: Path, tmp_path: Path) -> None:
    # Test 1: Directory path
    out_dir = tmp_path / "subdir"
    exit_code = main([
        str(synthetic_gdf),
        "--name=Planta1",
        f"-o={out_dir}/",
        "--quiet",
    ])
    assert exit_code == 0
    assert (out_dir / "survey_Planta1.xlsx").exists()
    assert (out_dir / "survey_Planta1.html").exists()

    # Test 2: File prefix
    out_prefix = tmp_path / "custom_prefix"
    exit_code_prefix = main([
        str(synthetic_gdf),
        f"-o={out_prefix}",
        "--quiet",
    ])
    assert exit_code_prefix == 0
    assert (tmp_path / "custom_prefix.xlsx").exists()
    assert (tmp_path / "custom_prefix.html").exists()


def test_invalid_gdf_fails_cli_and_raises_in_extractor(tmp_path: Path) -> None:
    bad_file = tmp_path / "corrupt.gdf"
    bad_file.write_bytes(b"not a valid cfbf stream at all")

    with pytest.raises(ValueError, match="Failed to parse GDF corrupt.gdf"):
        extract_gdf_survey(bad_file)

    exit_code = main([str(bad_file), "--no-excel", "--no-html", "--quiet"])
    assert exit_code != 0


def test_html_report_neutralizes_script_injection_and_handles_duplicate_sheets(tmp_path: Path) -> None:
    from gdf_survey.models import GdfSurveyResult, PumpRecord
    payload = '</script><script>globalThis.PWNED=1</script>'
    p1 = PumpRecord(1, "Equipo 1", "EQ_01", "CODE1", payload, "DEV1", "Tipo A", "Tipo A", "0", "0", "0", "0", "0", True, "1")
    p2 = PumpRecord(2, "Equipo 2", "EQ_02", "CODE2", "Bat2", "DEV2", "Tipo B", "Tipo B", "0", "0", "0", "0", "0", True, "1")

    r1 = GdfSurveyResult(Path("disp1.gdf"), "disp1", "DUPLICATE_NAME", "1", pumps=[p1])
    r2 = GdfSurveyResult(Path("disp2.gdf"), "disp2", "DUPLICATE_NAME", "1", pumps=[p2])

    out_html = tmp_path / "report.html"
    generate_html_survey([r1, r2], out_html)

    html_text = out_html.read_text(encoding="utf-8")
    # Payload must NOT contain literal unescaped closing script tag:
    assert payload not in html_text
    assert "\\u003c/script\\u003e" in html_text
    # Each screen must receive a distinct screen_id:
    assert "screen_0_disp1" in html_text
    assert "screen_1_disp2" in html_text


def test_excel_report_saves_formula_like_strings_as_literal_text(tmp_path: Path) -> None:
    from gdf_survey.models import GdfSurveyResult, PumpRecord
    formula_payload = "=1+1"
    hyperlink_payload = '=HYPERLINK("http://evil.com")'
    p = PumpRecord(1, "Equipo 1", formula_payload, hyperlink_payload, "Bat", "Dev", "Tipo A", "Tipo A", "0", "0", "0", "0", "0", True, "1")
    r = GdfSurveyResult(Path("test.gdf"), "test", "Sheet1", "1", pumps=[p])

    out_xlsx = tmp_path / "safe.xlsx"
    generate_excel_survey([r], out_xlsx)

    wb = openpyxl.load_workbook(out_xlsx, data_only=False)
    ws = wb["Sheet1"]
    cell_well = ws.cell(row=5, column=2)
    cell_code = ws.cell(row=5, column=3)

    assert cell_well.data_type == "s"
    assert cell_well.value == formula_payload
    assert cell_code.data_type == "s"
    assert cell_code.value == hyperlink_payload


def test_flat_mode_extraction(synthetic_gdf: Path) -> None:
    res = extract_gdf_survey(synthetic_gdf, layer_target="1", flat=True)
    assert res.total_items == 8
    assert res.total_objects == 8
    # Each item is individual object
    assert res.items[0].root_id == "eq1_disp"
    assert res.items[1].root_id == "eq1_well"


def test_custom_data_filter_extraction(synthetic_gdf: Path) -> None:
    res = extract_gdf_survey(
        synthetic_gdf,
        layer_target="1",
        custom_data_filter=["<<dispositivo>>", "<<pozo>>"],
    )
    assert res.discovered_custom_data_keys == ["<<dispositivo>>", "<<pozo>>"]
    item = res.items[0]
    assert "<<dispositivo>>" in item.custom_data
    assert "<<pozo>>" in item.custom_data
    assert "<<tienept>>" not in item.custom_data


def test_root_custom_data_and_pattern(synthetic_gdf: Path) -> None:
    res = extract_gdf_survey(
        synthetic_gdf,
        layer_target="1",
        root_name_pattern=r"^([a-z0-9]+)_",
        root_custom_data="<<dispositivo>>",
    )
    assert res.total_items == 1
    assert res.items[0].root_id == "eq1"
    assert res.items[0].primary_source == "EQ_01.CTRL_A"


def test_cli_generic_survey_options(synthetic_gdf: Path, tmp_path: Path) -> None:
    excel_out = tmp_path / "custom_cli.xlsx"
    html_out = tmp_path / "custom_cli.html"

    exit_code = main([
        str(synthetic_gdf),
        "-r=^([a-z0-9]+)_",
        "--custom-data=<<dispositivo>>,<<pozo>>",
        f"--out-excel={excel_out}",
        f"--out-html={html_out}",
        "--quiet",
    ])
    assert exit_code == 0
    assert excel_out.exists()
    assert html_out.exists()

    # Verify Excel headers contain only filtered custom data
    wb = openpyxl.load_workbook(excel_out)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=4, column=c).value for c in range(1, 8)]
    assert "<<dispositivo>>" in headers
    assert "<<pozo>>" in headers
    assert "<<tienept>>" not in headers


def test_cli_flat_mode_execution(synthetic_gdf: Path, tmp_path: Path) -> None:
    excel_out = tmp_path / "flat_cli.xlsx"
    html_out = tmp_path / "flat_cli.html"

    exit_code = main([
        str(synthetic_gdf),
        "--flat",
        f"--out-excel={excel_out}",
        f"--out-html={html_out}",
        "--quiet",
    ])
    assert exit_code == 0
    assert excel_out.exists()
    assert html_out.exists()


def test_generic_custom_data_html_report(tmp_path: Path) -> None:
    from gdf_survey.models import EquipmentRecord, GdfSurveyResult
    eq1 = EquipmentRecord(
        index=1,
        root_id="Motor_101",
        label="Motor 101",
        device_name="PLC_MAIN",
        controller_type="PLC",
        primary_source="[PLC]M101.Speed",
        custom_data={
            "<<tag>>": "M-101",
            "<<speed>>": "1450",
            "<<temperature>>": "58.5",
            "<<running>>": "1",
        },
    )
    res = GdfSurveyResult(
        Path("motor_disp.gdf"),
        "motor_disp",
        "Motors",
        "1",
        items=[eq1],
        discovered_custom_data_keys=["<<running>>", "<<speed>>", "<<tag>>", "<<temperature>>"],
    )

    out_html = tmp_path / "motors.html"
    generate_html_survey([res], out_html)
    assert out_html.exists()
    html_text = out_html.read_text(encoding="utf-8")
    assert "Motor_101" in html_text
    assert "\\u003c\\u003cspeed\\u003e\\u003e" in html_text or "speed" in html_text
    assert "PLC_MAIN" in html_text

